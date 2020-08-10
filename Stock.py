from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox
from datetime import date
from pandas import DataFrame
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import webbrowser

stock = {}
ing = {}
menu = {}
wastage = {}
old = {}
report = {}
meals = {}

f4 = ""

today = date.today()
month_inc = today.replace(month = today.month + 1)
months = month_inc.month
month_stock = today.replace(month = today.month + 2)
mstock = month_stock.month

works = ""
site = ""
pcode = {}

  
class Table: 
      
    def __init__(self,root):
        root.title("Report {}".format(today))
        print(ing)
          
        # code for creating table
        total_columns = 1
        v=0
        lst = []
        for key in report.keys():
            lst.append(key)
        total_rows = len(lst) +1
        for i in range(total_rows): 
            for j in range(total_columns):
                key = report
                t = lst[v]
                if i == 0:
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Product Code").grid(row = i, column = 1)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Ingredient").grid(row = i, column = 2)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Current Levels").grid(row = i, column = 3)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Wastage").grid(row = i, column = 4)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Used").grid(row = i, column = 5)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Predicted Levels").grid(row = i, column = 6)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Diffrence").grid(row = i, column = 7)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Unit").grid(row = i, column = 8)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Cost").grid(row = i, column = 9)
                else:                    
                    self.e = Label(root, width=25,relief = GROOVE, text = lst[v]).grid(row = i, column = 1)
                    self.e = Label(root, width=25,relief = GROOVE, text = report[t]["Ingredient"].capitalize()).grid(row = i, column = 2)
                    self.e = Label(root, width=25,relief = GROOVE, text = report[t]["Current Levels"]).grid(row = i, column = 3)
                    self.e = Label(root, width=25,relief = GROOVE, text = report[t]["Wastage"]).grid(row = i, column = 4)
                    self.e = Label(root, width=25,relief = GROOVE, text = report[t]["Used"]).grid(row = i, column = 5)
                    self.e = Label(root, width=25,relief = GROOVE, text = report[t]["Predicted Levels"]).grid(row = i, column = 6)
                    self.e = Label(root, width=25,relief = GROOVE, text = report[t]["Diffrence"]).grid(row = i, column = 7)
                    self.e = Label(root, width=25,relief = GROOVE, text = ing[t]["Unit"]).grid(row = i, column = 8)
                    self.e = Label(root, width=25,relief = GROOVE, text = "£{}".format(round(report[t]["Cost"],2))).grid(row = i, column = 9)
                    v= v +1

class StockTable: 
      
    def __init__(self,root):
        root.title("Stock {}".format(today))
        # code for creating table
        total_columns = 1
        v=0
        lst = []
        for key in ing.keys():
            lst.append(key)
        total_rows = len(lst) +1
        for i in range(total_rows): 
            for j in range(total_columns):
                key = report
                t = lst[v]
                if i == 0:
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Product Code").grid(row = i, column = 1)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Ingredient").grid(row = i, column = 2)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Count").grid(row = i, column = 3)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Unit").grid(row = i, column = 4)
                else:                    
                    self.e = Label(root, width=25,relief = GROOVE, text = lst[v]).grid(row = i, column = 1)
                    self.e = Label(root, width=25,relief = GROOVE, text = ing[t]["Ingredient"].capitalize()).grid(row = i, column = 2)
                    self.e = Entry(root, width=25,relief = GROOVE).grid(row = i, column = 3)
                    self.e = Label(root, width=25,relief = GROOVE, text = ing[t]["Unit"]).grid(row = i, column = 4)
                    v= v +1
        b = Button(root, width=20, text = "Printable Sheet", command = lambda: File_Creation()).grid( row = i+1, column = 2)
                    
class Meal_Table: 
      
    def __init__(self,root):
        root.title("Menu {}".format(today))
          
        # code for creating the menu table
        menu_dic()
        workbook = load_workbook(filename = works)
        menu_sheet = workbook["Menu"]
        total_columns = 1
        v=0
        lst = []
        for key in menu.keys():
            lst.append(key)
        total_rows = len(lst) +1
        for i in range(total_rows): 
            for j in range(total_columns):
                t = lst[v]
                print(t)
                m1 = menu[t]["Ingredient 1"]
                m2 = menu[t]["Ingredient 2"]
                m3 = menu[t]["Ingredient 3"]
                m4 = menu[t]["Ingredient 4"]
                if i == 0:
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Meal Code").grid(row = i, column = 1)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Meal Name").grid(row = i, column = 2)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Ingredient 1").grid(row = i, column = 3)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Ingredient 2").grid(row = i, column = 4)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Ingredient 3").grid(row = i, column = 5)
                    self.e = Label(root, width=20, relief = GROOVE, font = "bold", text = "Ingredient 4").grid(row = i, column = 6)
                else:                    
                    self.e = Label(root, width=25,relief = GROOVE, text = lst[v]).grid(row = i, column = 1)
                    self.e = Label(root, width=25,relief = GROOVE, text = menu[t]["Meal Name"]).grid(row = i, column = 2)
                    self.e = Label(root, width=25,relief = GROOVE, text = ing[m1]["Ingredient"].capitalize()).grid(row = i, column = 3)
                    self.e = Label(root, width=25,relief = GROOVE, text = ing[m2]["Ingredient"].capitalize()).grid(row = i, column = 4)
                    if menu[t]["Ingredient 3"] == None:
                        pass
                    else:
                        self.e = Label(root, width=25,relief = GROOVE, text = ing[m3]["Ingredient"].capitalize()).grid(row = i, column = 5)
                    if menu[t]["Ingredient 4"] == None:
                        pass
                    else:
                        self.e = Label(root, width=25,relief = GROOVE, text = ing[m4]["Ingredient"].capitalize()).grid(row = i, column = 6)
                    v= v +1
  
class App:
    def __init__(self, stock):
        stock.title("SWM Account")
        
        l1 = Label (stock, text = "Please enter the name of your site").grid(row = 1, column =1, columnspan =3)

        self.searched = StringVar(stock)
        name = Entry(stock, width = 50, borderwidth = 5,textvariable=self.searched).grid(row = 2, column =1, columnspan =3)
        submit = Button(stock, text = "Submit",height = "1", width = "15", command = lambda: self.name()).grid(row = 3, column =2)
        
    def name (self):
        global works
        global site
        works = self.searched.get() +".xlsx"
        site = self.searched.get()
        waste_dic()
        ing_dic()
        defult()

class Defult:
    def __init__(self, workspace):
        global f4
        workspace.title("{}".format(site))
        
        l1 = Label (workspace, text = "Welcome to {}".format(site)).grid(row = 0, column =0, columnspan =3)
        
        f1 = LabelFrame (workspace, text = "Stock Count", width = 3, bd = 5, padx = 5, pady = 5)
        f1.grid(row = 1, column = 0, columnspan =5)
        f2 = LabelFrame (workspace, text = "Wastage", width = 3, bd = 5, padx = 5, pady = 5)
        f2.grid(row = 1, column = 6, columnspan =5)
        f3 = LabelFrame (workspace, text = "Meal Plan", width = 3, bd = 5, padx = 5, pady = 5)
        f3.grid(row = 1, column = 11, columnspan =5)
        f4 = LabelFrame (workspace, text = "Report", width = 3, bd = 5, padx = 5, pady = 5)
        f4.grid(row = 6, column = 0, columnspan =20)

        #Stock count
        self.searched1 = StringVar(workspace)
        self.searched2 = StringVar(workspace)
        self.searched3 = StringVar(workspace)
        self.searched4 = StringVar(workspace)
        self.searched5 = StringVar(workspace)
        self.searched6 = StringVar(workspace)
        self.searched7 = StringVar(workspace)
        self.searched8 = StringVar(workspace)
        self.searched9 = StringVar(workspace)
        self.searched10 = StringVar(workspace)
        self.searched11 = StringVar(workspace)
        self.searched12 = StringVar(workspace)
        self.searched13 = StringVar(workspace)
        self.searched14 = StringVar(workspace)
        self.searched15 = StringVar(workspace)
        self.searched16 = StringVar(workspace)
        self.searched17 = StringVar(workspace)
        self.searched18 = StringVar(workspace)
        self.searched19 = StringVar(workspace)
        self.searched20 = StringVar(workspace)
        #Waastage
        self.searched21 = StringVar(workspace)
        self.searched22 = StringVar(workspace)
        self.searched23 = StringVar(workspace)
        self.searched24 = StringVar(workspace)
        self.searched25 = StringVar(workspace)
        self.searched26 = StringVar(workspace)
        self.searched27 = StringVar(workspace)
        self.searched28 = StringVar(workspace)
        self.searched29 = StringVar(workspace)
        self.searched30 = StringVar(workspace)
        self.searched31 = StringVar(workspace)
        self.searched32 = StringVar(workspace)
        self.searched33 = StringVar(workspace)
        self.searched34 = StringVar(workspace)
        self.searched35 = StringVar(workspace)
        self.searched36 = StringVar(workspace)
        self.searched37 = StringVar(workspace)
        self.searched38 = StringVar(workspace)
        self.searched39 = StringVar(workspace)
        self.searched40 = StringVar(workspace)

        #Meals
        self.searched41 = StringVar(workspace)
        self.searched42 = StringVar(workspace)
        self.searched43 = StringVar(workspace)
        self.searched44 = StringVar(workspace)
        self.searched45 = StringVar(workspace)
        self.searched46 = StringVar(workspace)
        self.searched47 = StringVar(workspace)
        self.searched48 = StringVar(workspace)
        self.searched49 = StringVar(workspace)
        self.searched50 = StringVar(workspace)
        self.searched51 = StringVar(workspace)
        self.searched52 = StringVar(workspace)
        self.searched53 = StringVar(workspace)
        self.searched54 = StringVar(workspace)
        self.searched55 = StringVar(workspace)
        self.searched56 = StringVar(workspace)
        self.searched57 = StringVar(workspace)
        self.searched58 = StringVar(workspace)
        self.searched59 = StringVar(workspace)
        self.searched60 = StringVar(workspace)
        
        #Stock Count
        l1 = Label (f1, text = "Product Code").grid(row = 0, column =0, columnspan =5)
        l2 = Label (f1, text = "Amount").grid(row = 0, column =6, columnspan =3)
        #l3 = Label (f1, text = "Pack Size").grid(row = 0, column = 9, columnspan =3)
        #Product code
        e1 = Entry(f1, width = 20, borderwidth = 5,textvariable=self.searched1).grid(row = 1, column =0, columnspan =5)
        e2 = Entry(f1, width = 20, borderwidth = 5,textvariable=self.searched2).grid(row = 2, column =0, columnspan =5)
        e3 = Entry(f1, width = 20, borderwidth = 5,textvariable=self.searched3).grid(row = 3, column =0, columnspan =5)
        e4 = Entry(f1, width = 20, borderwidth = 5,textvariable=self.searched4).grid(row = 4, column =0, columnspan =5)
        e5 = Entry(f1, width = 20, borderwidth = 5,textvariable=self.searched5).grid(row = 5, column =0, columnspan =5)
        e6 = Entry(f1, width = 20, borderwidth = 5,textvariable=self.searched6).grid(row = 6, column =0, columnspan =5)
        e7 = Entry(f1, width = 20, borderwidth = 5,textvariable=self.searched7).grid(row = 7, column =0, columnspan =5)
        e8 = Entry(f1, width = 20, borderwidth = 5,textvariable=self.searched8).grid(row = 8, column =0, columnspan =5)
        e9 = Entry(f1, width = 20, borderwidth = 5,textvariable=self.searched9).grid(row = 9, column =0, columnspan =5)
        e10 = Entry(f1, width = 20, borderwidth = 5,textvariable=self.searched10).grid(row = 10, column =0, columnspan =5)
        #Amounts
        e11 = Entry(f1, width = 10, borderwidth = 5,textvariable=self.searched11).grid(row = 1, column =6, columnspan =5)
        e12 = Entry(f1, width = 10, borderwidth = 5,textvariable=self.searched12).grid(row = 2, column =6, columnspan =5)
        e13 = Entry(f1, width = 10, borderwidth = 5,textvariable=self.searched13).grid(row = 3, column =6, columnspan =5)
        e14 = Entry(f1, width = 10, borderwidth = 5,textvariable=self.searched14).grid(row = 4, column =6, columnspan =5)
        e15 = Entry(f1, width = 10, borderwidth = 5,textvariable=self.searched15).grid(row = 5, column =6, columnspan =5)
        e16 = Entry(f1, width = 10, borderwidth = 5,textvariable=self.searched16).grid(row = 6, column =6, columnspan =5)
        e17 = Entry(f1, width = 10, borderwidth = 5,textvariable=self.searched17).grid(row = 7, column =6, columnspan =5)
        e18 = Entry(f1, width = 10, borderwidth = 5,textvariable=self.searched18).grid(row = 8, column =6, columnspan =5)
        e19 = Entry(f1, width = 10, borderwidth = 5,textvariable=self.searched19).grid(row = 9, column =6, columnspan =5)
        e20 = Entry(f1, width = 10, borderwidth = 5,textvariable=self.searched20).grid(row = 10, column =6, columnspan =5)

        submit = Button(f1, text = "Submit",height = "1", width = "15", command = lambda: self.dicto()).grid(row = 11, column =3)

        #Wastage
        l3 = Label (f2, text = "Product Code").grid(row = 0, column =0, columnspan =5)
        l4 = Label (f2, text = "Amount").grid(row = 0, column =6, columnspan =3)
        w1 = Entry(f2, width = 20, borderwidth = 5,textvariable=self.searched21).grid(row = 1, column =0, columnspan =5)
        w2 = Entry(f2, width = 20, borderwidth = 5,textvariable=self.searched22).grid(row = 2, column =0, columnspan =5)
        w3 = Entry(f2, width = 20, borderwidth = 5,textvariable=self.searched23).grid(row = 3, column =0, columnspan =5)
        w4 = Entry(f2, width = 20, borderwidth = 5,textvariable=self.searched24).grid(row = 4, column =0, columnspan =5)
        w5 = Entry(f2, width = 20, borderwidth = 5,textvariable=self.searched25).grid(row = 5, column =0, columnspan =5)
        w6 = Entry(f2, width = 20, borderwidth = 5,textvariable=self.searched26).grid(row = 6, column =0, columnspan =5)
        w7 = Entry(f2, width = 20, borderwidth = 5,textvariable=self.searched27).grid(row = 7, column =0, columnspan =5)
        w8 = Entry(f2, width = 20, borderwidth = 5,textvariable=self.searched28).grid(row = 8, column =0, columnspan =5)
        w9 = Entry(f2, width = 20, borderwidth = 5,textvariable=self.searched29).grid(row = 9, column =0, columnspan =5)
        w10 = Entry(f2, width = 20, borderwidth = 5,textvariable=self.searched30).grid(row = 10, column =0, columnspan =5)

        w11 = Entry(f2, width = 10, borderwidth = 5,textvariable=self.searched31).grid(row = 1, column =6, columnspan =5)
        w12 = Entry(f2, width = 10, borderwidth = 5,textvariable=self.searched32).grid(row = 2, column =6, columnspan =5)
        w13 = Entry(f2, width = 10, borderwidth = 5,textvariable=self.searched33).grid(row = 3, column =6, columnspan =5)
        w14 = Entry(f2, width = 10, borderwidth = 5,textvariable=self.searched34).grid(row = 4, column =6, columnspan =5)
        w15 = Entry(f2, width = 10, borderwidth = 5,textvariable=self.searched35).grid(row = 5, column =6, columnspan =5)
        w16 = Entry(f2, width = 10, borderwidth = 5,textvariable=self.searched36).grid(row = 6, column =6, columnspan =5)
        w17 = Entry(f2, width = 10, borderwidth = 5,textvariable=self.searched37).grid(row = 7, column =6, columnspan =5)
        w18 = Entry(f2, width = 10, borderwidth = 5,textvariable=self.searched38).grid(row = 8, column =6, columnspan =5)
        w19 = Entry(f2, width = 10, borderwidth = 5,textvariable=self.searched39).grid(row = 9, column =6, columnspan =5)
        w20 = Entry(f2, width = 10, borderwidth = 5,textvariable=self.searched40).grid(row = 10, column =6, columnspan =5)

        submit = Button(f2, text = "Submit",height = "1", width = "15", command = lambda: self.wdicto()).grid(row = 11, column =3)

        #Meal Plan
        l1 = Label (f3, text = "Meal Code").grid(row = 0, column =0, columnspan =5)
        l2 = Label (f3, text = "Amount Cooked").grid(row = 0, column =6, columnspan =3)
        
        m1 = Entry(f3, width = 20, borderwidth = 5,textvariable=self.searched41).grid(row = 1, column =0, columnspan =5)
        m2 = Entry(f3, width = 20, borderwidth = 5,textvariable=self.searched42).grid(row = 2, column =0, columnspan =5)
        m3 = Entry(f3, width = 20, borderwidth = 5,textvariable=self.searched43).grid(row = 3, column =0, columnspan =5)
        m4 = Entry(f3, width = 20, borderwidth = 5,textvariable=self.searched44).grid(row = 4, column =0, columnspan =5)
        m5 = Entry(f3, width = 20, borderwidth = 5,textvariable=self.searched45).grid(row = 5, column =0, columnspan =5)
        m6 = Entry(f3, width = 20, borderwidth = 5,textvariable=self.searched46).grid(row = 6, column =0, columnspan =5)
        m7 = Entry(f3, width = 20, borderwidth = 5,textvariable=self.searched47).grid(row = 7, column =0, columnspan =5)
        m8 = Entry(f3, width = 20, borderwidth = 5,textvariable=self.searched48).grid(row = 8, column =0, columnspan =5)
        m9 = Entry(f3, width = 20, borderwidth = 5,textvariable=self.searched49).grid(row = 9, column =0, columnspan =5)
        m10 = Entry(f3, width = 20, borderwidth = 5,textvariable=self.searched50).grid(row = 10, column =0, columnspan =5)

        m11 = Entry(f3, width = 10, borderwidth = 5,textvariable=self.searched51).grid(row = 1, column =6, columnspan =5)
        m12 = Entry(f3, width = 10, borderwidth = 5,textvariable=self.searched52).grid(row = 2, column =6, columnspan =5)
        m13 = Entry(f3, width = 10, borderwidth = 5,textvariable=self.searched53).grid(row = 3, column =6, columnspan =5)
        m14 = Entry(f3, width = 10, borderwidth = 5,textvariable=self.searched54).grid(row = 4, column =6, columnspan =5)
        m15 = Entry(f3, width = 10, borderwidth = 5,textvariable=self.searched55).grid(row = 5, column =6, columnspan =5)
        m16 = Entry(f3, width = 10, borderwidth = 5,textvariable=self.searched56).grid(row = 6, column =6, columnspan =5)
        m17 = Entry(f3, width = 10, borderwidth = 5,textvariable=self.searched57).grid(row = 7, column =6, columnspan =5)
        m18 = Entry(f3, width = 10, borderwidth = 5,textvariable=self.searched58).grid(row = 8, column =6, columnspan =5)
        m19 = Entry(f3, width = 10, borderwidth = 5,textvariable=self.searched59).grid(row = 9, column =6, columnspan =5)
        m20 = Entry(f3, width = 10, borderwidth = 5,textvariable=self.searched60).grid(row = 10, column =6, columnspan =5)

        submit = Button(f3, text = "Submit",height = "1", width = "15", command = lambda: self.mdicto()).grid(row = 11, column =3)

        #Report
        submit = Button(f4, text = "Generate Report", command = lambda: Report()).grid(row = 1, column =1)
        submit = Button(f4, text = "Meal List", command = lambda: Table_meal()).grid(row = 1, column =3)
        submit = Button(f4, text = "Stock List", command = lambda: Table_stock()).grid(row = 1, column =5)
        submit = Button(f4, text = "Year to Date Graph", command = lambda: YTD()).grid(row = 1, column =6)

    def mdicto(self): #dictonary of meals
        global pcode
        pcode = {self.searched41.get():self.searched51.get(),
                 self.searched42.get():self.searched52.get(),
                 self.searched43.get():self.searched53.get(),
                 self.searched44.get():self.searched54.get(),
                 self.searched45.get():self.searched55.get(),
                 self.searched46.get():self.searched56.get(),
                 self.searched47.get():self.searched57.get(),
                 self.searched48.get():self.searched58.get(),
                 self.searched49.get():self.searched59.get(),
                 self.searched50.get():self.searched60.get()
                 }
        Meal_Record()
        
    def wdicto(self): #dictonary of wastage
        global pcode
        pcode = {self.searched21.get():self.searched31.get(),
                 self.searched22.get():self.searched32.get(),
                 self.searched23.get():self.searched33.get(),
                 self.searched24.get():self.searched34.get(),
                 self.searched25.get():self.searched35.get(),
                 self.searched26.get():self.searched36.get(),
                 self.searched27.get():self.searched37.get(),
                 self.searched28.get():self.searched38.get(),
                 self.searched29.get():self.searched39.get(),
                 self.searched30.get():self.searched40.get()
                 }
        waste()

    def dicto(self): #dictionary of stock
        global pcode
        pcode = {self.searched1.get():self.searched11.get(),
                 self.searched2.get():self.searched12.get(),
                 self.searched3.get():self.searched13.get(),
                 self.searched4.get():self.searched14.get(),
                 self.searched5.get():self.searched15.get(),
                 self.searched6.get():self.searched16.get(),
                 self.searched7.get():self.searched17.get(),
                 self.searched8.get():self.searched18.get(),
                 self.searched9.get():self.searched19.get(),
                 self.searched10.get():self.searched20.get()
                 }
        stock_add()
    
        


def stock_add (): #Full stock count
    workbookname = works
    workbook = load_workbook(filename = workbookname)
    ingredients_sheet = workbook["Ingredients"]
    sl = workbook.sheetnames
    if "Stock {}".format(today) not in sl:
        workbook.create_sheet("Stock {}".format(today),0)
        stock_sheet = workbook["Stock {}".format(today)]
        stock_sheet["A1"] = "Product Code"
        stock_sheet["B1"] = "Ingredients"
        stock_sheet["C1"] = "Amount"
        stock_sheet["D1"] = "Pack Size"
        stock_sheet["E1"] = "Cost"
    else:
        pass
    stock_sheet = workbook["Stock {}".format(today)]
    r = 1
    l=list(pcode.keys())
    for i in l:
        if i == "":
            pass
        elif i in ing:
            row = stock_sheet.max_row
            r= int(row) + 1
            f = pcode[i]
            a = "A{}".format(r)
            b = "B{}".format(r)
            c = "C{}".format(r)
            d = "D{}".format(r)
            e = "E{}".format(r)
            stock_sheet[a] = i
            stock_sheet[b] = ing[i]['Ingredient']
            stock_sheet[c] = f
            stock_sheet[d] = ing[i]['Pack size']
            stock_sheet[e] = round(float(ing[i]['Cost']) * (float(f)/int(ing[i]['Pack size'])),2)
            stock_sheet[e].number_format = '[$£-809]#,##0.00;[RED]-[$£-809]#,##0.00'
            workbook.save(filename = workbookname)
        else:
            messagebox.showerror("Error","{} is not in my database".format(i))
    if "Stock" in sl[2]:
        v = sl[2]
        workbook.remove_sheet(v)
    else:
        pass
    workbook.save(filename = workbookname)
    com()
    

def waste (): #Enters the wastage
    workbookname = works
    workbook = load_workbook(filename = workbookname)
    wastage_sheet = workbook["Wastage"]
    row = wastage_sheet.max_row
    cleanup()
    r = int(row) + 1
    l = []
    for w in wastage_sheet["A"]:
        l.append(w.value)
    for i in pcode:
        if i == "":
            pass
        elif i not in ing:
             messagebox.showerror("Error","{} is not in my database".format(i))
        elif i in l:
            rs = 1
            for rows in wastage_sheet["A"]:
                if rows.value == i:
                    c = "C{}".format(rs)
                    d = "D{}".format(rs)
                    wastage_sheet[c] = int(wastage_sheet[c].value) + int(pcode[i])
                    wastage_sheet[d] = round(float(ing[i]['Cost']) * int(wastage_sheet[c].value)/int(ing[i]['Pack size']),2)
                    wastage_sheet[d].number_format = '[$£-809]#,##0.00;[RED]-[$£-809]#,##0.00'
                elif rows.value != i:
                    rs = rs + 1
            workbook.save(filename = workbookname)
                
        else:
            f = pcode[i]
            a = "A{}".format(r)
            a = "A{}".format(r)
            b = "B{}".format(r)
            c = "C{}".format(r)
            d = "D{}".format(r)
            e = "E{}".format(r)
            F = "F{}".format(r)
            wastage_sheet[a] = i
            wastage_sheet[b] = ing[i]['Ingredient']
            wastage_sheet[c] = f
            wastage_sheet[d] = round(float(ing[i]['Cost']) * (float(f)/int(ing[i]['Pack size'])),2)
            wastage_sheet[d].number_format = '[$£-809]#,##0.00;[RED]-[$£-809]#,##0.00'
            wastage_sheet[e] = today
            wastage_sheet[F] = months
            workbook.save(filename = workbookname)
            r = r + 1
    workbook.save(filename = workbookname)
    com()
    
def cleanup (): # deletes wastage older than 1 month
    workbookname = works
    workbook = load_workbook(filename = workbookname)
    wastage_sheet = workbook["Wastage"]
    row = 1
    for d in wastage_sheet["F"]:
        f = str(d.value)
        if f == None:
            pass
        else:
            if f == "":
                break
            elif str(f) not in str(today.month):
                row = row +1
            elif str(f) in str(today.month):
                wastage_sheet.delete_rows(row, 1)
            else:
                worng()
                break
    workbook.save(filename = workbookname)
            

        
def stock_dic (): #dictonry of stock
    workbook = load_workbook(filename = works)
    sheets = workbook.sheetnames
    stock_sheet = workbook[sheets[0]]
    global stock
    pr =0
    for p in stock_sheet["A"]:
        if p.value == "":
            break
        else:
            pr = pr +1
    for row in stock_sheet.iter_rows(min_row = 2,
                                   max_row = pr,
                                   min_col = 1,
                                   max_col = 5,
                                   ):
        product_code = row[0].value
        product = {
            "Ingredient" : row[1].value,
            "Amount": row[2].value,
            "Pack Size" : row[3].value,
            "Cost" : row[4].value
            }
        stock[product_code] = product
        
def old_dic (): #dictonry of old stock
    workbook = load_workbook(filename = works)
    sheets = workbook.sheetnames
    stock_sheet = workbook[sheets[1]]
    global old
    pr =0
    for p in stock_sheet["A"]:
        if p.value == "":
            break
        else:
            pr = pr +1
    for row in stock_sheet.iter_rows(min_row = 2,
                                   max_row = pr,
                                   min_col = 1,
                                   max_col = 5,
                                   ):
        product_code = row[0].value
        product = {
            "Ingredient" : row[1].value,
            "Amount": row[2].value,
            "Pack Size" : row[3].value,
            "Cost" : row[4].value
            }
        old[product_code] = product

def meals_dic (): #dictonry of meals cooked
    workbook = load_workbook(filename = works)
    meal_sheet = workbook["Meals"]
    global meals
    pr =0
    for p in meal_sheet["A"]:
        if p.value == "":
            break
        else:
            pr = pr +1
    for row in meal_sheet.iter_rows(min_row = 2,
                                   max_row = pr,
                                   min_col = 1,
                                   max_col = 2,
                                   ):
        product_code = row[0].value
        product = {
            "Amount" : row[1].value
            }
        meals[product_code] = product

def tables_dic (): #dictonry of the table data
    workbook = load_workbook(filename = works)
    report_sheet = workbook["Report"]
    global report
    pr =0
    for p in report_sheet["A"]:
        if p.value == "":
            break
        else:
            pr = pr +1
    for row in report_sheet.iter_rows(min_row = 2,
                                   max_row = pr,
                                   min_col = 1,
                                   max_col = 8,
                                   ):
        product_code = row[0].value
        product = {
            "Ingredient" : row[1].value,
            "Current Levels": row[2].value,
            "Wastage" : row[3].value,
            "Used" : row[4].value,
            "Predicted Levels" : row[5].value,
            "Diffrence" : row[6].value,
            "Cost" : row[7].value
            }
        report[product_code] = product

def ing_dic (): #Indredients list in a dictionary
    workbook = load_workbook(filename = works)
    ingredients_sheet = workbook["Ingredients"]
    global ing
    pr =0
    for p in ingredients_sheet["A"]:
        if p.value == "":
            break
        else:
            pr = pr +1
    for row in ingredients_sheet.iter_rows(min_row = 2,
                                   max_row = pr,
                                   min_col = 1,
                                   max_col = 5,
                                   ):
        product_code = row[0].value
        product = {
            "Ingredient" : row[1].value,
            "Pack size": row[2].value,
            "Unit" : row[3].value,
            "Cost" : row[4].value
            }
        ing[product_code] = product

def menu_dic (): #dictonary of the menu
    workbook = load_workbook(filename = works)
    menu_sheet = workbook["Menu"]
    global menu
    pr =0
    for p in menu_sheet["A"]:
        if p.value == "":
            break
        else:
            pr = pr +1
    for row in menu_sheet.iter_rows(min_row = 2,
                                   max_row = pr,
                                   min_col = 1,
                                   max_col = 10,
                                   ):
        meal = row[0].value
        product = {
            "Meal Name" : row[1].value,
            "Ingredient 1" : row[2].value,
            "Amount 1":row[3].value,
            "Ingredient 2": row[4].value,
            "Amount 2":row[5].value,
            "Ingredient 3" : row[6].value,
            "Amount 3":row[7].value,
            "Ingredient 4" : row[8].value,
            "Amount 4":row[9].value
            }
        menu[meal] = product

def waste_dic (): #dictonary of the wastage
    workbook = load_workbook(filename = works)
    wastage_sheet = workbook["Wastage"]
    global wastage
    pr =0
    for p in wastage_sheet["A"]:
        if p.value == "":
            break
        else:
            pr = pr +1
    for row in wastage_sheet.iter_rows(min_row = 2,
                                   max_row = pr,
                                   min_col = 1,
                                   max_col = 9,
                                   ):
        product_code = row[0].value
        product = {
            "Ingredient" : row[1].value,
            "Amount" : row[2].value,
            "Cost" : row[3].value
            }
        wastage[product_code] = product
        
def Report (): #Runs the report 
    workbook = load_workbook(filename = works)
    report_sheet = workbook["Report"]
    meals_sheet = workbook["Meals"]
    wastage_sheet = workbook["Wastage"]
    sheets = workbook.sheetnames
    stock_sheet = workbook[sheets[0]]
    old_sheet = workbook[sheets[1]]
    menu_dic()
    meals_dic()
    old_dic()
    stock_dic()
    cleanup()
    waste_dic()
    r1 = 2
    r2 = 1
    r3 = 2
    l = list(ing.keys())
    l2 = wastage
    l3 = list(stock.keys())
    l4 = list(menu.keys())
    l5 = list(old.keys())
    for i in l: #resets the sheet
        report_sheet["A{}".format(r1)] = i
        report_sheet["B{}".format(r1)] = ing[i]["Ingredient"]
        report_sheet["C{}".format(r1)] = "0"
        report_sheet["D{}".format(r1)] = "0"
        report_sheet["E{}".format(r1)] = "0"
        report_sheet["F{}".format(r1)] = "0"
        report_sheet["G{}".format(r1)] = "0"
        report_sheet["H{}".format(r1)] = "0"
        report_sheet["H{}".format(r1)].number_format = '[$£-809]#,##0.00;[RED]-[$£-809]#,##0.00'
        workbook.save(filename = works)
        r1 = r1 + 1
    for y in l2: #Transfers the wastage over
        r2 = 1
        for row in report_sheet["A"]:
            v = row.value
            if y == v:
                report_sheet["D{}".format(r2)] = int(wastage[y]["Amount"])
            elif y != v:
                r2 = r2 + 1
            else:
                worng()
    for u in l3: #trasfers the stock over
        r3 = 1
        for row in report_sheet["A"]:
            v = row.value
            if u == v:
                report_sheet["C{}".format(r3)] = int(stock[u]["Amount"])
            elif u != v:
                r3 = r3 + 1
            else:
                worng()
    for row in meals_sheet["A"]: #Collects the meals cooked
        v1 = row.value
        if v1 in l4: #makes sure they are in the dict of meals
            a = 1
            for i in menu[v1]: #Gets the Ingrediants
                r4 = 1
                for rows in report_sheet["A"]:
                    vr =rows.value
                    if menu[v1][i] == vr:
                        k = float(meals[v1]["Amount"])
                        s = report_sheet["E{}".format(r4)] = int(report_sheet["E{}".format(r4)].value)+ (float(menu[v1]["Amount {}".format(a)]* k))
                        a = a + 1
                    else:
                        r4 = r4 +1
        else:
            pass
    r6 = 1
    for u in l5: #trasfers the old stock over
        r5 = 1
        r6 =r6 +1
        for row in report_sheet["A"]:
            v = row.value
            if u == v:
                report_sheet["F{}".format(r5)] = int(old[u]["Amount"]) - int(report_sheet["D{}".format(r5)].value) - int(report_sheet["E{}".format(r5)].value)
            elif u != v and v not in l5:
                if r5 == 1:
                    pass
                else:
                    report_sheet["F{}".format(r5)] = int(report_sheet["D{}".format(r5)].value) - int(report_sheet["E{}".format(r5)].value)
                r5 = r5 + 1
            elif u != v:
                r5 = r5 +1
            else:
                worng()
    r7 = 1
    for row in report_sheet["A"]: #works out the diffrence in stock
        if r7 == 1:
            pass
        else:
            h =int(report_sheet["C{}".format(r7)].value) -  int(report_sheet["F{}".format(r7)].value)
            report_sheet["G{}".format(r7)] = h
        r7 = r7 + 1
    r8 = 1
    for row in report_sheet["A"]: #works out the cost diffrence in stock
        v = row.value
        if r8 == 1:
            pass
        else:
            h =int(report_sheet["C{}".format(r8)].value) -  int(report_sheet["F{}".format(r8)].value)
            report_sheet["H{}".format(r8)] = int(h)/ int(ing[v]["Pack size"]) * float(ing[v]["Cost"])
        r8 = r8 + 1
    workbook.save(filename = works)
    tables_dic()
    Table_report()

def Meal_Record():
    menu_dic()
    workbookname = works
    workbook = load_workbook(filename = workbookname)
    meal_sheet = workbook["Meals"]
    row = meal_sheet.max_row
    l=list(pcode.keys())
    for i in l:
        if i == "":
            pass
        elif i in menu:
            row = meal_sheet.max_row
            r= int(row) + 1
            f = pcode[i]
            a = "A{}".format(r)
            b = "B{}".format(r)
            c = "C{}".format(r)
            d = "D{}".format(r)
            meal_sheet[a] = i
            meal_sheet[b] = f
            meal_sheet[c] = today
            meal_sheet[d] = months
            workbook.save(filename = workbookname)
        else:
            messagebox.showerror("Error","{} is not in my database".format(i))
        Meal_clean()


def Meal_clean():
    workbookname = works
    workbook = load_workbook(filename = workbookname)
    meal_sheet = workbook["Meals"]
    row = 1
    for d in meal_sheet["D"]:
        f = str(d.value)
        if f == None:
            pass
        else:
            if f == "":
                break
            elif str(f) not in str(today.month):
                row = row +1
            elif str(f) in str(today.month):
                meal_sheet.delete_rows(row, 1)
                workbook.save(filename = workbookname)
            else:
                worng()
                break
    workbook.save(filename = workbookname)

def YTD():
    workbookname = works
    workbook = load_workbook(filename = workbookname)
    ytd_sheet = workbook["YTD Report"]
    wastage_sheet = workbook["Wastage"]
    sheets = workbook.sheetnames
    stock_sheet = workbook[sheets[0]]
    svalues = 0
    wvalues = 0
    r = ytd_sheet.max_row + 1
    for row in stock_sheet["E"]:
        t = row.value
        if t == None:
            break
        elif t == "":
            break
        elif t == "Cost":
            pass
        else:
            svalues = svalues + int(t)
    for row in wastage_sheet["D"]:
        t = row.value
        if t == None:
            break
        elif t == "":
            break
        elif t == "Cost":
            pass
        else:
            wvalues = wvalues + int(t)
    for row in ytd_sheet:
        ytd_sheet["A{}".format(r)] = today
        ytd_sheet["B{}".format(r)] = int(svalues) - int(wvalues)
        ytd_sheet["B{}".format(r)].number_format = '[$£-809]#,##0.00;[RED]-[$£-809]#,##0.00'
    i = ytd_sheet.max_row
    graph_dic = {}
    for row in ytd_sheet.iter_rows(min_row = 2,
                                   max_row = i,
                                   min_col = 1,
                                   max_col = 2
                                   ):
        date = row[0].value
        total = row[1].value
        graph_dic[date] = total

    workbook.save(filename = workbookname)
    Graph()

def Graph():
    workbookname = works
    workbook = load_workbook(filename = workbookname)
    ytd_sheet = workbook["YTD Report"]
    totals = []
    dates = []
    data = {}
    
    for row in ytd_sheet["B"]:
        if row.value == "Total":
            pass
        elif row.value == None:
            break
        else:
            totals.append(row.value)
            
    for row in ytd_sheet["A"]:
        if row.value == "Date":
            pass
        elif row.value == None:
            break
        else:
            dates.append(row.value)
    data["Cost Totals"] = totals
    data["Date"] = dates
    df = DataFrame(data, columns = ["Date", "Cost Totals"])
    
    
    Toplevel = Tk()
    Toplevel.title("Stock Loss Report")

    figure2 = plt.Figure(figsize=(20,20), dpi=100)
    ax2 = figure2.add_subplot(111)
    line2 = FigureCanvasTkAgg(figure2, Toplevel)
    line2.get_tk_widget().pack(side=LEFT, fill=BOTH)
    df2 = df[['Date','Cost Totals']].groupby('Date').sum()
    df2.plot(kind='line', legend=False, ax=ax2, color='Black',marker='o', fontsize=10, grid = True)
    ax2.set_title('Stock Cost Report')
    
def File_Creation():
    ing_dic ()
    file1 = open("Count_Sheet.txt", "w")
    
    for i in ing:
        file1.write(i)
        file1.write("|")
        file1.write(ing[i]["Ingredient"].capitalize())
        l = len(ing[i]["Ingredient"])
        l2 = len(ing[i]["Unit"])
        if l < 10:
            for v in range(10-l):
                file1.write(" ")
        file1.write("|")
        file1.write("             ")
        file1.write("|")
        file1.write(ing[i]["Unit"])
        if l2 < 6:
            for v in range(6-l2):
                file1.write(" ")
        file1.write("|")
        file1.write("\n")
        file1.write("___")
        file1.write("|")
        file1.write("__________")
        file1.write("|")
        file1.write("_____________")
        file1.write("|")
        file1.write("______|")
        file1.write("\n")
    file1.close()
    webbrowser.open("Count_Sheet.txt")
    
def wrong():
     messagebox.showerror("Error", "Something went wrong")

def com():
     messagebox.showinfo("Title", "Job Complete")
    

def startup(): # runs the start up
    root = Tk()
    app = App(root)
    root.mainloop()

def defult(): # The workspace
    root = Tk()
    app = Defult(root)
    root.mainloop()

def Table_report(): # The report table
    root = Tk()
    app = Table(root)
    root.mainloop()

def Table_meal(): # The menu list
    root = Tk()
    app = Meal_Table(root)
    root.mainloop()

def Table_stock(): # The menu list
    root = Tk()
    app = StockTable(root)
    root.mainloop()
