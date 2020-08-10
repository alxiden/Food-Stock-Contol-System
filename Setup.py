from openpyxl import *
from tkinter import *

works = ""

class App:
    def __init__(self, window):
        window.title("SWM Account Maker")
        
        l1 = Label (window, text = "Please choose a name for the account, you will be asked for this on login").grid(row = 1, column =1, columnspan =3)

        self.searched = StringVar(window)
        name = Entry(window, width = 50, borderwidth = 5,textvariable=self.searched).grid(row = 2, column =1, columnspan =3)
        submit = Button(window, text = "Submit", command = lambda: self.name()).grid(row = 3, column =2)

    def name (self):
        global works
        works = self.searched.get()
        setup()
        
    
def setup ():
    #sitename = input("If you are sure you want to make a new account please enter the name again: ").strip().lower()
    #workbookname = sitename + ".xlsx"
    workbookname = works + ".xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Stock"
    wastage = workbook.create_sheet("Wastage")
    report = workbook.create_sheet("Report")
    meals_s = workbook.create_sheet("Meals Served")
    meals_b = workbook.create_sheet("Meal Breakdown")
    ingrediants= workbook.create_sheet("Ingredients")
    stock_sheet = workbook["Stock"]
    stock_sheet["A1"] = "Product Code"
    stock_sheet["B1"] = "Ingredients"
    stock_sheet["C1"] = "Amount"
    stock_sheet["D1"] = "Pack Size"
    stock_sheet["E1"] = "Cost"
    stock_sheet["E2"].number_format = '[$£-809]#,##0.00;[RED]-[$£-809]#,##0.00'
    
    wastage_sheet = workbook["Wastage"]
    wastage_sheet["A1"] = "Product Code"
    wastage_sheet["B1"] = "Ingredients"
    wastage_sheet["C1"] = "Amount"
    wastage_sheet["D1"] = "Cost"
    wastage_sheet["E1"] = "Initials"
    wastage_sheet["D2"].number_format = '[$£-809]#,##0.00;[RED]-[$£-809]#,##0.00'

    report_sheet = workbook["Report"]

    meal_s_sheet = workbook["Meals Served"]
    meal_s_sheet["A1"] = "Sunday"
    meal_s_sheet["B1"] = "Monday"
    meal_s_sheet["C1"] = "Tuesday"
    meal_s_sheet["D1"] = "Wednesday"
    meal_s_sheet["E1"] = "Thursday"
    meal_s_sheet["F1"] = "Friday"
    meal_s_sheet["G1"] = "Saturday"

    meal_b_sheet = workbook["Meal Breakdown"]

    ingrediants_sheet = workbook["Ingredients"]
    ingrediants_sheet["A1"] = "Product Code"
    ingrediants_sheet["B1"] = "Ingrediants"
    ingrediants_sheet["C1"] = "Pack Size"
    ingrediants_sheet["D1"] = "Pack Weight"
    ingrediants_sheet["E1"] = "Cost"
    ingrediants_sheet["E2"].number_format = '[$£-809]#,##0.00;[RED]-[$£-809]#,##0.00'

    workbook.save(filename = workbookname)

def setep():
    root = Tk()
    app = App(root)
    root.mainloop()

